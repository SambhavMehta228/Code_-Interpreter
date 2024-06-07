import PyPDF2  # For reading PDFs
import openpyxl  # For reading Excel files
import pandas as pd  # For reading CSVs
from docx import Document  # For reading Word documents
import openai  # For using GPT-3.5
import logging  # For logging stuff

# Setting up logging so we know what's happening
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# OpenAI API key, replace with your actual key
openai.api_key = "YOUR_OPENAI_API_KEY"

# Function to read PDF files
def read_pdf(file_path):
    logging.info(f"Reading PDF file: {file_path}")
    content = ""
    try:
        with open(file_path, "rb") as file:
            reader = PyPDF2.PdfFileReader(file)
            for page_num in range(reader.numPages):
                content += reader.getPage(page_num).extractText()
    except Exception as e:
        logging.error(f"Failed to read PDF file: {e}")
        return None
    return content

# Function to read Excel files
def read_xlsx(file_path):
    logging.info(f"Reading XLSX file: {file_path}")
    content = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_content = [[cell.value for cell in row] for row in sheet.iter_rows()]
            content.append({sheet_name: sheet_content})
    except Exception as e:
        logging.error(f"Failed to read XLSX file: {e}")
        return None
    return content

# Function to read CSV files
def read_csv(file_path):
    logging.info(f"Reading CSV file: {file_path}")
    try:
        df = pd.read_csv(file_path)
        return df.to_string()
    except Exception as e:
        logging.error(f"Failed to read CSV file: {e}")
        return None

# Function to read Word documents
def read_docx(file_path):
    logging.info(f"Reading DOCX file: {file_path}")
    content = ""
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            content += para.text + "\n"
    except Exception as e:
        logging.error(f"Failed to read DOCX file: {e}")
        return None
    return content

# Function to generate code using GPT-3.5
def generate_code(file_content, user_prompt):
    logging.info("Generating Python code using GPT-3.5 API")
    prompt = f"Based on the following content:\n{file_content}\n\nAnd the user's request:\n{user_prompt}\n\nGenerate the relevant Python code:"
    try:
        response = openai.Completion.create(
            engine="text-davinci-003",
            prompt=prompt,
            max_tokens=150,
            temperature=0.5,
        )
        code = response.choices[0].text.strip()
    except Exception as e:
        logging.error(f"Failed to generate code: {e}")
        return None
    return code

# Function to execute the generated code
def execute_code(code):
    logging.info("Executing generated Python code")
    local_vars = {}
    try:
        exec(code, {"__builtins__": None}, local_vars)
        return local_vars
    except Exception as e:
        logging.error(f"Code execution failed: {e}")
        return str(e)

# Main function to process the file and do everything
def process_file(file_path, file_type, user_prompt):
    logging.info(f"Processing file: {file_path} of type {file_type}")
    content = None
    if file_type == "pdf":
        content = read_pdf(file_path)
    elif file_type == "xlsx":
        content = read_xlsx(file_path)
    elif file_type == "csv":
        content = read_csv(file_path)
    elif file_type == "docx":
        content = read_docx(file_path)
    else:
        return "Unsupported file type"
    
    if content is None:
        return "Failed to read file content"

    code = generate_code(content, user_prompt)
    if code is None:
        return "Failed to generate code"

    output = execute_code(code)
    return output

# Function to make the output look nice
def format_output(output):
    logging.info("Formatting output for user")
    if isinstance(output, dict):
        return "\n".join([f"{key}: {value}" for key, value in output.items()])
    return output

# Example usage, this is where the magic happens
if __name__ == "__main__":
    file_path = "example.pdf"  # Change this to your file
    file_type = "pdf"  # Change this to your file type
    user_prompt = "Summarize the main points"  # Change this to your prompt

    output = process_file(file_path, file_type, user_prompt)
    formatted_output = format_output(output)
    print(formatted_output)
